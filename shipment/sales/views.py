from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from .forms import ExcelUploadForm
from .models import Sale
import openpyxl
import uuid
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib.auth import authenticate
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required






@login_required
def sales_home(request):
    message = None
    duplicate_serials_in_file = []
    duplicate_serials_in_db = []
    search_query = request.GET.get('search_serial', '')

    # Fetch only the last 10 sales records
    sales = Sale.objects.order_by('-id')[:10]

    # Filter sales by serial number if a search query is provided
    if search_query:
        sales = Sale.objects.filter(serial_lot__icontains=search_query)

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            try:
                # Get all existing serial numbers from the database
                serial_numbers_in_db = {str(s).strip().lower() for s in Sale.objects.values_list('serial_lot', flat=True)}
                serial_numbers_in_file = set()

                # Process the file and check for duplicates
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) < 24:  # Ensure the row has enough columns
                        continue

                    serial_lot = str(row[0]).strip().lower()  # Normalize serial number

                    # Skip duplicate checks for serial numbers starting with "LOT"
                    if serial_lot.startswith("lot"):
                        continue

                    # Check for duplicates within the file
                    if serial_lot in serial_numbers_in_file:
                        duplicate_serials_in_file.append(serial_lot)
                    else:
                        serial_numbers_in_file.add(serial_lot)

                # Debugging (optional): Log serial numbers for verification
                print("Serial numbers in file:", serial_numbers_in_file)

                # Check for duplicates in the database, excluding "LOT" serial numbers
                duplicate_serials_in_db = [
                    serial for serial in serial_numbers_in_file
                    if serial in serial_numbers_in_db and not serial.startswith("lot")
                ]

                # Combine errors for file and database duplicates
                if duplicate_serials_in_file or duplicate_serials_in_db:
                    error_messages = []
                    if duplicate_serials_in_file:
                        error_messages.append(
                            f"The file contains duplicate serial numbers: {', '.join(set(duplicate_serials_in_file))}"
                        )
                    if duplicate_serials_in_db:
                        error_messages.append(
                            f"The following serial numbers already exist in the database: {', '.join(set(duplicate_serials_in_db))}"
                        )
                    message = "Upload aborted! " + " ".join(error_messages)
                    return render(request, 'sales/sales.html', {
                        'form': form,
                        'message': message,
                        'sales': sales,
                        'duplicate_serials': duplicate_serials_in_file + duplicate_serials_in_db,
                        'search_query': search_query,
                    })

                # If no duplicates, save the data to the database
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) < 24:  # Skip rows with insufficient columns
                        continue

                    try:
                        invoice_date = datetime.strptime(row[8], '%d.%m.%Y').date() if row[8] else None
                        posting_date = datetime.strptime(row[18], '%d.%m.%Y').date() if row[18] else None
                    except ValueError:
                        invoice_date = None
                        posting_date = None

                    Sale.objects.create(
                        serial_lot=row[0],
                        document_number=row[1],
                        item_number=row[2],
                        item_reference_number=row[3],
                        quantity=row[4],
                        unit_of_measure_code=row[5],
                        box=row[6],
                        invoice_number=row[7],
                        invoice_date=invoice_date,
                        shipment_number=row[9],
                        description=row[10],
                        carbon_qty=row[11],
                        lot_carbon=row[12],
                        sales_order_number=row[13],
                        sales_order_line_number=row[14],
                        customer_order_number=row[15],
                        customer_order_line_number=row[16],
                        sales_amount_actual=row[17],
                        posting_date=posting_date,
                        project_code=row[19],
                        hs_code=row[20],
                        net_weight_kg=row[21],
                        type=row[22],
                        lu_number=row[23],
                    )
                message = "Data uploaded successfully!"
            except Exception as e:
                message = f"Error processing file: {e}"
        else:
            message = "Invalid file format. Please upload a valid Excel file."
    else:
        form = ExcelUploadForm()

    return render(request, 'sales/sales.html', {
        'form': form,
        'message': message,
        'sales': sales,
        'duplicate_serials': duplicate_serials_in_file + duplicate_serials_in_db,
        'search_query': search_query,
    })






@login_required
def shelf(request):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"

    # Add headers to the Excel sheet (including "LU Number")
    headers = [
        "Serial/Lot", "Document Number", "Item Number", "Item Reference No", "Quantity",
        "Unit of Measure Code", "Box", "Invoice Number", "Invoice Date", "Shipment No",
        "Description", "Carbon Qty", "Lot Carbon", "Sales Order No", "Sales Order Line No",
        "Customer Order No", "Customer Order Line No", "Sales Amount (Actual)", "Posting Date",
        "Project Code", "HS Code", "Net Weight (KG)", "Type", "Batch Number", "LU Number"  # Added LU Number
    ]
    ws.append(headers)

    # Fetch data from the database and populate the sheet
    for sale in Sale.objects.all():
        ws.append([
            sale.serial_lot,
            sale.document_number,
            sale.item_number,
            sale.item_reference_number,
            sale.quantity,
            sale.unit_of_measure_code,
            sale.box,
            sale.invoice_number,
            sale.invoice_date.isoformat() if sale.invoice_date else None,
            sale.shipment_number,
            sale.description,
            sale.carbon_qty,
            sale.lot_carbon,
            sale.sales_order_number,
            sale.sales_order_line_number,
            sale.customer_order_number,
            sale.customer_order_line_number,
            sale.sales_amount_actual,
            sale.posting_date.isoformat() if sale.posting_date else None,
            sale.project_code,
            sale.hs_code,
            sale.net_weight_kg,
            sale.type,
            str(sale.batch_number),  # Batch Number
            sale.lu_number,  # Added LU Number
        ])

    # Create a response for downloading the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sales_data.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response








@login_required
def delete_sales_database(request):
    if request.method == 'POST':
        admin_password = request.POST.get('admin_password')

        # Authenticate the admin password
        admin_user = authenticate(username='admin', password=admin_password)
        if admin_user is not None:
            # Delete all records in the Sale table
            Sale.objects.all().delete()
            messages.success(request, "Database deleted successfully!")
            return redirect('sales_home')
        else:
            messages.error(request, "Invalid admin password. Database was not deleted.")

    # Render the delete confirmation page
    return render(request, 'sales/delete_database.html')




def delete_line(request, pk):
    # Get the Sale object or raise a 404 error if not found
    sale = get_object_or_404(Sale, pk=pk)
    sale.delete()  # Delete the Sale record
    return HttpResponseRedirect(reverse('sales_home'))  # Redirect to the sales home page





from django.contrib.auth import authenticate

def delete_by_file(request):
    if request.method == 'POST':
        admin_password = request.POST.get('admin_password')  # Get admin password from the form
        file = request.FILES.get('file')  # Get the uploaded file

        if not admin_password:
            messages.error(request, "Admin password is required.")
            return redirect('sales_home')

        # Authenticate admin password
        admin_user = authenticate(username='admin', password=admin_password)
        if admin_user is None:
            messages.error(request, "Invalid admin password. Operation aborted.")
            return redirect('sales_home')

        if not file:
            messages.error(request, "No file uploaded.")
            return redirect('sales_home')

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            # Check for header row with "Serial/Lot" and "Batch Number"
            headers = [cell.value for cell in sheet[1]]
            if "Serial/Lot" not in headers or "Batch Number" not in headers:
                messages.error(request, "Invalid file format. Please upload a valid file with 'Serial/Lot' and 'Batch Number'.")
                return redirect('sales_home')

            # Fetch indices of the required columns
            serial_index = headers.index("Serial/Lot")
            batch_index = headers.index("Batch Number")

            deleted_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                serial_lot = str(row[serial_index]).strip()
                batch_number = str(row[batch_index]).strip()

                # Delete matching records in the database
                deleted_rows = Sale.objects.filter(serial_lot=serial_lot, batch_number=batch_number).delete()
                deleted_count += deleted_rows[0]  # Count the number of rows deleted

            messages.success(request, f"Successfully deleted {deleted_count} records.")
        except Exception as e:
            messages.error(request, f"Error processing file: {e}")

        return redirect('sales_home')
    else:
        return render(request, 'sales/delete_by_file.html')